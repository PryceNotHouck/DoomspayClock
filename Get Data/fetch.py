# Used here: documentation and GeeksForGeeks pages associated with requests, XLRD, and OpenPyXL

import zipfile
import requests
import csv
import openpyxl
import xlrd
import os
from datetime import datetime, timedelta


def flush_local():
    dir_path = os.path.join(os.path.dirname(__file__), "Local")
    if os.path.exists(dir_path):
        files = os.listdir(dir_path)
        for file in files:
            file_path = os.path.join(dir_path, file)
            if os.path.isfile(file_path):
                os.remove(file_path)


def convert_xls(file):
    file_path = "Local/" + str(file)
    csv_file_path = os.path.splitext(file_path)[0] + ".csv"
    workbook = xlrd.open_workbook(file_path)
    sheet = workbook.sheet_by_index(0)

    # Accounts for Excel's old base date (1900-01-01)
    base_date = datetime(1899, 12, 30)

    with open(csv_file_path, 'w', newline='') as csvfile:
        col = csv.writer(csvfile)
        for row in range(sheet.nrows):
            row_values = sheet.row_values(row)
            if isinstance(row_values[0], float):
                date_value = base_date + timedelta(days=row_values[0])
                row_values[0] = date_value.strftime("%Y-%m-%d")
            col.writerow(row_values)
    os.remove(file_path)


def convert_xlsx(file):
    file_path = "Local/" + str(file)
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    csv_file_path = os.path.splitext(file_path)[0] + ".csv"
    with open(csv_file_path, 'w', newline="") as csvfile:
        col = csv.writer(csvfile)
        for row in sheet.iter_rows(values_only=True):
            col.writerow(row)
    os.remove(file_path)


def convert_xls_bea(file):
    file_path = file
    csv_file_path = os.path.splitext(file_path)[0] + ".csv"
    workbook = xlrd.open_workbook(file_path)

    # Accounts for Excel's old base date (1900-01-01)
    base_date = datetime(1899, 12, 30)

    with open(csv_file_path, 'w', newline='') as csvfile:
        col = csv.writer(csvfile)

        for sheet_index in range(workbook.nsheets):
            sheet = workbook.sheet_by_index(sheet_index)
            for row in range(sheet.nrows):
                row_values = sheet.row_values(row)
                if isinstance(row_values[0], float) and sheet.cell_type(row, 0) == xlrd.XL_CELL_DATE:
                    date_value = xlrd.xldate.xldate_as_datetime(row_values[0], workbook.datemode)
                    row_values[0] = date_value.strftime("%Y-%m-%d")
                col.writerow(row_values)
            if sheet_index < workbook.nsheets - 1:
                col.writerow([])

    os.remove(file_path)


def convert_xlsx_bea(file):
    try:
        workbook = openpyxl.load_workbook(file)
        csv_file_path = os.path.splitext(file)[0] + ".csv"

        with open(csv_file_path, 'w', newline="") as csvfile:
            col = csv.writer(csvfile)

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows(values_only=True):
                    col.writerow(row)
                col.writerow([])
        os.remove(file)
    except zipfile.BadZipfile:
        pass


def fetch_bea():
    print("Warning: This takes several minutes to complete.")
    base_url = "https://raw.githubusercontent.com/PryceNotHouck/DoomspayClock/master/Datasets/BEA%20GDP/"
    base_path = "Local/BEA/"

    if not os.path.exists(base_path):
        os.makedirs(base_path)

    for year in range(6, 25):
        year_str = f"200{year}" if year <= 9 else f"20{year}"
        year_url = f"{base_url}{year_str}/"
        year_path = os.path.join(base_path, year_str)

        if not os.path.exists(year_path):
            os.makedirs(year_path)

        quarter_range = 2 if year_str == "2024" else 5

        for quarter in range(1, quarter_range):
            quarter_url = f"{year_url}Q{quarter}/"
            quarter_path = os.path.join(year_path, f"Q{quarter}")

            if not os.path.exists(quarter_path):
                os.makedirs(quarter_path)

            if int(year_str) <= 2009:
                section_range = 9
            else:
                section_range = 8

            for section in range(0, section_range):
                if int(year_str) <= 2010:
                    file_url = f"{quarter_url}Section{section}ALL_xls.xls"
                    file_path = os.path.join(quarter_path, f"Section{section}ALL_xls.xls")
                elif int(year_str) >= 2017:
                    if int(year_str) == 2017 and quarter < 3:
                        file_url = f"{quarter_url}Section{section}all_xls.xls"
                        file_path = os.path.join(quarter_path, f"Section{section}all_xls.xls")
                    else:
                        file_url = f"{quarter_url}Section{section}all_xls.xlsx"
                        file_path = os.path.join(quarter_path, f"Section{section}all_xls.xlsx")
                else:
                    file_url = f"{quarter_url}Section{section}all_xls.xls"
                    file_path = os.path.join(quarter_path, f"Section{section}all_xls.xls")
                print("Fetching:", file_path)
                res = requests.get(file_url, allow_redirects=True)
                if res.status_code == 200:
                    with open(file_path, 'wb') as f:
                        f.write(res.content)
                    if file_path.endswith(".xls"):
                        convert_xls_bea(file_path)
                    else:
                        convert_xlsx_bea(file_path)
                else:
                    print(f"Failed to fetch {file_url}, retrying.")
                    file_url = f"{quarter_url}Section{section}all_xls.xls"
                    file_path = os.path.join(quarter_path, f"Section{section}all_xls.xls")
                    print("Fetching:", file_path)
                    res = requests.get(file_url, allow_redirects=True)
                    if res.status_code == 200:
                        with open(file_path, 'wb') as f:
                            f.write(res.content)
                        if file_path.endswith(".xls"):
                            convert_xls_bea(file_path)
                        else:
                            convert_xlsx_bea(file_path)
                    else:
                        print(f"Failed to fetch {file_url}, cannot retry.")


def fetch(var, file):
    print("Fetching:", file)
    url = "https://raw.githubusercontent.com/PryceNotHouck/DoomspayClock/master/Datasets/"
    match var:
        case 0:
            # GDP Growth
            fetch_bea()
        case 1:
            # Inflation Rate, CPI
            url += "Statista/" + file
        case 2:
            # Trade/Foreign Investment
            fetch_bea()
        case 3:
            # Unemployment
            url += "BLS/" + file
        case 4:
            # Housing Prices
            url += "FHFA/" + file
        case 5:
            # Loan Default Rates, Credit Delinquency
            url += "FRED/" + file
        case 6:
            # Cboe
            url += "Cboe/" + file

    res = requests.get(url, allow_redirects=True)
    file_path = "Local/" + file
    if file.endswith(".xlsx"):
        with open(file_path, 'wb') as f:
            f.write(res.content)
        convert_xlsx(file)
    elif file.endswith(".xls"):
        with open(file_path, 'wb') as f:
            f.write(res.content)
        convert_xls(file)
    else:
        with open(file_path, 'wb') as f:
            f.write(res.content)


def get_unemployment():
    year = -1
    for i in range(0, 24):
        year += 1
        file = "laucnty" + "{:02d}".format(year) + ".xlsx"
        fetch(3, file)
    year = 89
    for i in range(0, 10):
        year += 1
        file = "laucnty" + "{:02d}".format(year) + ".xlsx"
        fetch(3, file)


def get_inflation():
    fetch(1, "inflationData.xlsx")


def get_cpi():
    fetch(1, "cpi_data.xlsx")


def get_housing():
    fetch(4, "housingprices.xls")


def get_cboe():
    fetch(6, "GVZ_History.csv")
    fetch(6, "OVX_History.csv")
    fetch(6, "VIXS&P9D_History.csv")
    fetch(6, "VVIX_History.csv")
    fetch(6, "VXAPL_History.csv")
    fetch(6, "VXAZN_History.csv")
    fetch(6, "VXEEM_History.csv")


def get_loan_default():
    fetch(5, "DRBLACBS.xls")
    fetch(5, "DRALACBN.xls")


def get_credit():
    fetch(5, "DRALACBN.xls")
    fetch(5, "DRCLACBS.xls")
    fetch(5, "DRCCLACBS.xls")


if __name__ == "__main__":
    fetch_bea()
